from signal import CTRL_C_EVENT
from tkinter import END, Button, Entry, IntVar, Radiobutton, Tk , Event
from tkinter.filedialog import askdirectory, askopenfilename
from tkinter.messagebox import showerror, showinfo, showwarning
from pandas import ExcelFile
from openpyxl import load_workbook
from re import split
from os import getpid, kill, listdir # operating system === windows = linux
from string import Template

def getInfoStudents(path: str):
    names , emails , percentages , grades = [] , [] , [] , []
    excelFile = ExcelFile(path)
    file = load_workbook(path)
    sheetName = excelFile.parse(file.sheetnames[0])
    number_rows = sheetName.shape[0]
    
    for i in range(0 , number_rows - 1):
        names.append(sheetName["Name"][i])
        emails.append(sheetName["Email"][i])
        percentages.append(sheetName["Percentage"][i])
        grades.append(sheetName["Grade"][i])
        
    return names , emails , percentages , grades

def get_attachments(path: str):
    
    def digit(text: str):
        return int(text) if text.isdigit() else text
    
    def keys(text: str):
        return [digit(i) for i in split(r"(\d+)" , text)] # d: integer 
    
    pdf_list = []
    for pdf in listdir(path):
        pdf_list.append(path + "/" +  pdf)   
          
    pdf_list.sort(key= keys)
    return pdf_list

def read_message(message_filepath : str):
    try:
        with open(message_filepath , "r" , encoding= "UTF-8") as messagefile:
            msgContent = messagefile.read()
        return Template(msgContent)
    except:
        print("Error No such file or directory")

def display_connection_problems():
    showwarning(title = "Potential Connection Problems!",
                message = "Kindly Check your Internet Connection...")

def display_wrong_credentials():
    showinfo(title = "Invalid Credentials!" , message = "wrong username or password")

def display_path_check():
    showerror(title="Invalid Path!", message="Kindly Check your Choice and the Provided Path...")

def display_successfull_send():
    showinfo(title = "Congrats" , message = "All Emails ha been sent successfully")

def display_unexpected_error():
    showerror(title = "Unexpected Error!!" , message = "Something Went Wrong")
    
def display_empty_fields():
    showerror(title="Required Field!", message="Mandatory Data is Missing...")

def browse_excel_file(excel: Entry):
    filename = askopenfilename(initialdir = "C:" , title = "Please select the excel file")
    excel.config(bg = "white")
    excel.delete(0 , END)
    excel.insert(0 , filename)
    
def browse_template_file(msg: Entry):
    filename = askopenfilename(initialdir = "C:" , title = "Please select the message template file")
    msg.config(bg = "white")
    msg.delete(0 , END)
    msg.insert(0 , filename)

def yesButton(ent1 : Entry , ent2: Entry , radio1: Radiobutton , radio2: Radiobutton ,
              yes: Button , no: Button , browse:Button , send: Button , entries : list[Entry]):
    for entry in entries:
        entry.delete(0, END)
    yes.config(bg = "green")
    no.config(bg = "black")
    radio1.config(state= "normal")
    radio2.config(state= "normal")
    ent1.config(state = "normal")
    ent2.config(state = "normal")
    browse.config(state = "disable")
    send.config(state= "normal")
    
def noButton(ent1 : Entry , ent2: Entry , radio1: Radiobutton , radio2: Radiobutton ,
              yes: Button , no: Button , browse:Button , send: Button , entries : list[Entry]):
    for entry in entries:
        entry.delete(0, END)
    yes.config(bg = "black")
    no.config(bg = "red")
    radio1.config(state= "disable")
    radio2.config(state= "disable")
    ent1.config(state = "disable")
    ent2.config(state = "disable")
    browse.config(state = "disable")
    send.config(state= "normal")
    
def toggleRadioButton(choice : IntVar , radio1: Radiobutton , radio2 : Radiobutton , btn:Button):
    if choice.get() == 1:
        radio1.config(selectcolor= "green")
        radio2.config(selectcolor= "red") 
    elif choice.get() == 2:
        radio1.config(selectcolor= "red")
        radio2.config(selectcolor= "green")
    btn.config(state = "normal")
    
def browse_path_files(choice: IntVar ,browse : Button , path:Entry):
    if choice.get() == 1 or choice.get() == 2:
            global f
            if choice.get() == 1:
                def oneAttachment():
                    filename = askopenfilename(initialdir= "C:", 
                                               title = "Please Select the Attachment file")
                    return filename
                f = oneAttachment()
        
            elif choice.get() == 2:
                def multipleAttachment():
                    filename = askdirectory(initialdir= "C:", 
                                               title = "Please Select the Attachment file directory")
                    return filename
                f = multipleAttachment()
                
            path.config(bg= "white")
            path.delete(0 , END)
            path.insert(0 , f)
    else:
        browse.config(state = "disable")

def printout(name : str):
    print (f"Email Sent to {name}!")
    print (f"Email(s) Sent!")
    print ("*" * 50)
        
def interrupt_app():
    kill(getpid(), CTRL_C_EVENT) # send CTRL + C to app (only in windows)

def handle_entry_focus_red(event : Event):
    if event.widget["bg"] == "red":
        event.widget["bg"] = "white"

def handle_button_focus_red(event : Event):
    if event.widget["textvariable"] == "":
        event.widget["bg"] = "white"

def check_entries(root: Tk , entries_list : list[Entry]):
    boolean = True
    for entry in entries_list:
        if entry["state"] == "normal" and not entry.get():
            root.focus()
            entry.config(bg="red")
            entry.bind()
            boolean = False
            
    if boolean == False:
        display_empty_fields()
    
def get_widgets_config(widgets_list : list):
    widgets_config_keys = []
    widgets_config_items = []
    
    for widget in widgets_list:
        single_widget_config_keys = []
        single_widget_config_items = []
        
        for key in widget.keys():
            single_widget_config_keys.append(key)
            single_widget_config_items.append(widget[key])
            
        widgets_config_keys.append(single_widget_config_keys)
        widgets_config_items.append(single_widget_config_items)
        
    return widgets_config_keys, widgets_config_items

def reset_widgets_config(widgets_list, widgets_config_keys, widgets_config_items):
    i = 0
    for widget in widgets_list:
        if type(widget) == type(Entry()):
            widget.delete(0, "end")
        k = 0
        for key in widgets_config_keys[i]:
            widget[key] = widgets_config_items[i][k]
            k = k + 1
        i = i + 1

def reset_all(root: Tk ,choice : IntVar , btn : Button , entries_list, entries_config_keys,
              entries_config_items, buttons_list, buttons_config_keys, buttons_config_items):
    root.focus()
    choice.set(0)
    btn.invoke()
    reset_widgets_config(entries_list, entries_config_keys, entries_config_items)
    reset_widgets_config(buttons_list, buttons_config_keys, buttons_config_items)
    root.update_idletasks()